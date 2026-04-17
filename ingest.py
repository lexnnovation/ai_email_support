import os
import fitz
import openpyxl
from app.core.config import SUPABASE_URL, SUPABASE_KEY, OPENAI_API_KEY
from app.services.embedding_service import get_embedding
from app.db.supabase_client import supabase

EMBEDDING_DIR = "embedding_files"
XLSX_FILE = "Summer Online Support - Estivo Supporto Clienti online Airbnb - Booking .xlsx"

# Sheets to process as general
GENERAL_SHEETS = ["FAQ ENG", "Pre Sales FAQ - Eng",
                  "FAQ ITA", "info base", "Domande"]
# Sheets to skip
SKIP_SHEETS = ["Contatti utili", "Appartamenti con problemi"]

# Unique PDFs (ignore numbered duplicates)
UNIQUE_PDFS = {
    "Augustus Apartment Manual.pdf": "Augustus",
    "Manual Lcc 6.pdf": "Lcc6",
    "Manual Maredena-English Version.pdf": "Maredena",
    "Manuale Bafile 18.pdf": "Bafile 18",
    "Manuale Delta.pdf": "Delta",
    "Manuale Euroresidence-English.pdf": "Euroresidence",
    "Manuale I Bagni 36.pdf": "I Bagni 36",
    "Manuale La Torre.pdf": "La Torre",
    "Manuale Maria Luisa.pdf": "MariaLuisa",
    "Manuale Quito.pdf": "Quito",
    "Manuale San Silvestro.pdf": "San Silvestro",
    "Manuale SunnyHome C2 & C3.pdf": "SunnyHome C2 & C3",
    "Manuale SunnyHome C8.pdf": "SunnyHome C8",
    "Manuale Villa.pdf": "Villa",
    "Manuale Volta.pdf": "Volta",
}

CHUNK_SIZE = 1500
CHUNK_OVERLAP = 200
MIN_CHUNK_LENGTH = 50


def chunk_by_paragraphs(text, max_size=CHUNK_SIZE, min_size=MIN_CHUNK_LENGTH):
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    chunks = []
    current = ""
    for line in lines:
        if len(current) + len(line) + 1 <= max_size:
            current = (current + " " + line).strip()
        else:
            if current and len(current) >= min_size:
                chunks.append(current)
            current = line
    if current and len(current) >= min_size:
        chunks.append(current)
    return chunks


def insert_document(content, embedding, apartment, doc_type="faq"):
    supabase.table("documents").insert({
        "content": content,
        "embedding": embedding,
        "apartment": apartment,
        "type": doc_type,
    }).execute()


def ingest_pdf(filepath, apartment):
    print(
        f"  Processing PDF: {os.path.basename(filepath)} -> apartment={apartment}")
    doc = fitz.open(filepath)
    full_text = ""
    for page in doc:
        full_text += page.get_text()
    doc.close()

    chunks = chunk_by_paragraphs(full_text)
    print(f"    Extracted {len(chunks)} chunks")
    for i, chunk in enumerate(chunks):
        embedding = get_embedding(chunk)
        insert_document(chunk, embedding, apartment, "policy")
        print(f"    Inserted chunk {i+1}/{len(chunks)}")


def ingest_xlsx_sheet(wb, sheet_name, apartment):
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    print(
        f"  Processing sheet: {sheet_name} -> apartment={apartment} ({len(rows)} rows)")

    inserted = 0
    for row in rows:
        cells = [str(c).strip()
                 for c in row if c is not None and str(c).strip()]
        if not cells:
            continue

        combined = " ".join(cells)
        if len(combined) < MIN_CHUNK_LENGTH:
            continue

        # For Q&A sheets, format as "Question: ... Answer: ..."
        if len(cells) >= 2:
            content = f"Question: {cells[0]} Answer: {cells[1]}"
        else:
            content = combined

        # Chunk if content is too long
        if len(content) > CHUNK_SIZE:
            chunks = chunk_text(content)
        else:
            chunks = [content]

        for chunk in chunks:
            embedding = get_embedding(chunk)
            insert_document(chunk, embedding, apartment, "faq")
            inserted += 1

    print(f"    Inserted {inserted} documents")


def ingest_pdfs_only():
    print("=" * 60)
    print("PDF RE-INGESTION (paragraph chunking, type=policy)")
    print("=" * 60)
    pdf_apartments = list(UNIQUE_PDFS.values())
    print(
        f"Deleting existing PDF docs for {len(pdf_apartments)} apartments...")
    for apt in pdf_apartments:
        supabase.table("documents").delete().eq(
            "apartment", apt).eq("type", "faq").execute()
        supabase.table("documents").delete().eq(
            "apartment", apt).eq("type", "policy").execute()
    print("Deleted. Starting re-ingestion...")
    for filename, apartment in UNIQUE_PDFS.items():
        filepath = os.path.join(EMBEDDING_DIR, filename)
        if os.path.exists(filepath):
            ingest_pdf(filepath, apartment)
        else:
            print(f"  WARNING: File not found: {filename}")
    print("\n" + "=" * 60)
    print("PDF RE-INGESTION COMPLETE")
    print("=" * 60)


def main():
    print("=" * 60)
    print("DATA INGESTION START")
    print("=" * 60)

    # --- INGEST PDFs ---
    print("\n--- PDF INGESTION ---")
    for filename, apartment in UNIQUE_PDFS.items():
        filepath = os.path.join(EMBEDDING_DIR, filename)
        if os.path.exists(filepath):
            ingest_pdf(filepath, apartment)
        else:
            print(f"  WARNING: File not found: {filename}")

    # --- INGEST XLSX ---
    print("\n--- XLSX INGESTION ---")
    xlsx_path = os.path.join(EMBEDDING_DIR, XLSX_FILE)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            print(f"  Skipping sheet: {sheet_name}")
            continue

        if sheet_name in GENERAL_SHEETS:
            apartment = "general"
        else:
            apartment = sheet_name

        ingest_xlsx_sheet(wb, sheet_name, apartment)

    wb.close()

    print("\n" + "=" * 60)
    print("DATA INGESTION COMPLETE")
    print("=" * 60)


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "--pdfs-only":
        ingest_pdfs_only()
    else:
        main()
